"""
Generate Hoski Web Desk campaign build Excel file.
Run: python3 build_campaign_excel.py
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "webdesk-campaign-build.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
C_DARK       = "1A1A2E"   # deep navy  – sheet tab / heavy headers
C_MID        = "16213E"   # mid navy   – section headers
C_ACCENT     = "0F3460"   # blue       – sub-headers
C_HIGHLIGHT  = "E94560"   # red-pink   – flags / warnings
C_LIGHT_BLUE = "D6E4F0"   # pale blue  – alternating row
C_WHITE      = "FFFFFF"
C_LIGHT_GREY = "F5F5F5"
C_YELLOW     = "FFF9C4"   # warning bg
C_GREEN      = "E8F5E9"   # positive / included
C_RED_LIGHT  = "FFEBEE"   # excluded

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_header_row(ws, row, values, bg=C_ACCENT, fg=C_WHITE,
                     bold=True, size=10):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg, size=size)
        c.alignment = center()
        c.border = border_thin()

def write_data_row(ws, row, values, bg=C_WHITE, bold=False,
                   font_color=C_DARK, wrap=True):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=font_color)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=wrap)
        c.border = border_thin()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def merge_header(ws, row, text, col_start, col_end,
                 bg=C_MID, fg=C_WHITE, size=11):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=size)
    c.alignment = center()
    c.border = border_thin()

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ── Data ─────────────────────────────────────────────────────────────────────

CAMPAIGN_SETTINGS = [
    ["Setting", "hoski_webdesign_search_1apr26", "hoski_cro_search_1apr26"],
    ["Campaign Type", "Search", "Search"],
    ["Daily Budget (CAD)", "$50", "$50"],
    ["Monthly Budget Est.", "~$1,500", "~$1,500"],
    ["Bid Strategy (Launch)", "Maximize Clicks", "Maximize Clicks"],
    ["Bid Strategy (Post 30 conv.)", "Maximize Conversions", "Maximize Conversions"],
    ["Networks", "Search only — Display Network OFF", "Search only — Display Network OFF"],
    ["Search Partners", "ON (review after 30 days)", "ON (review after 30 days)"],
    ["Geo Target", "Canada", "Canada"],
    ["City Bid Adjustments", "+15% Montreal, Toronto, Vancouver, Calgary", "+15% Montreal, Toronto, Vancouver, Calgary"],
    ["Language", "English + French", "English + French"],
    ["Ad Schedule", "All hours (review after 30 days)", "All hours (review after 30 days)"],
    ["Match Type Strategy", "Exact + Phrase — NO Broad (new account)", "Exact + Phrase — NO Broad (new account)"],
    ["Conversion Goal", "Strategy Call Booked", "Strategy Call Booked"],
    ["Landing Page", "https://hoski.ca/web-desk", "https://hoski.ca/web-desk"],
    ["Ad Rotation", "Optimize", "Optimize"],
    ["URL Tracking Template", "?utm_source=google&utm_medium=cpc&utm_campaign={campaignid}&utm_content={adgroupid}&utm_term={keyword}", "?utm_source=google&utm_medium=cpc&utm_campaign={campaignid}&utm_content={adgroupid}&utm_term={keyword}"],
]

# keywords: (ad_group, keyword, match_type, intent, notes)
WD_KEYWORDS = [
    # webdesign_agency
    ("webdesign_agency", "web design agency", "Exact", "Transactional", "Core term"),
    ("webdesign_agency", "website design company", "Exact", "Transactional", "Core term"),
    ("webdesign_agency", "web design firm", "Exact", "Transactional", "Core term"),
    ("webdesign_agency", "website design agency", "Exact", "Transactional", "Core term"),
    ("webdesign_agency", "web design company canada", "Exact", "Transactional + Geo", ""),
    ("webdesign_agency", "website design company canada", "Exact", "Transactional + Geo", ""),
    ("webdesign_agency", "web design agency", "Phrase", "Transactional", ""),
    ("webdesign_agency", "website design company", "Phrase", "Transactional", ""),
    ("webdesign_agency", "web design firm canada", "Phrase", "Transactional + Geo", ""),
    ("webdesign_agency", "professional web design", "Phrase", "Commercial", ""),
    ("webdesign_agency", "website design services", "Phrase", "Commercial", ""),
    ("webdesign_agency", "web agency canada", "Phrase", "Transactional + Geo", ""),
    ("webdesign_agency", "website design agency canada", "Phrase", "Transactional + Geo", ""),
    ("webdesign_agency", "hire web design agency", "Phrase", "Transactional", "High intent"),
    # webdesign_redesign
    ("webdesign_redesign", "website redesign", "Exact", "Transactional", "Core term"),
    ("webdesign_redesign", "website redesign service", "Exact", "Transactional", ""),
    ("webdesign_redesign", "website revamp", "Exact", "Transactional", ""),
    ("webdesign_redesign", "redesign my website", "Exact", "Transactional", "High intent"),
    ("webdesign_redesign", "website overhaul", "Exact", "Transactional", ""),
    ("webdesign_redesign", "website makeover", "Exact", "Commercial", ""),
    ("webdesign_redesign", "website refresh service", "Exact", "Commercial", ""),
    ("webdesign_redesign", "site redesign agency", "Exact", "Transactional", ""),
    ("webdesign_redesign", "website redesign", "Phrase", "Transactional", ""),
    ("webdesign_redesign", "website redesign service", "Phrase", "Transactional", ""),
    ("webdesign_redesign", "website revamp service", "Phrase", "Transactional", ""),
    ("webdesign_redesign", "redesign my website", "Phrase", "Transactional", ""),
    ("webdesign_redesign", "website makeover agency", "Phrase", "Transactional", ""),
    ("webdesign_redesign", "professional website redesign", "Phrase", "Transactional", ""),
    # webdesign_convert
    ("webdesign_convert", "web design that converts", "Exact", "Transactional", "Highest intent"),
    ("webdesign_convert", "website that generates leads", "Exact", "Transactional", "Highest intent"),
    ("webdesign_convert", "conversion focused web design", "Exact", "Transactional", ""),
    ("webdesign_convert", "high converting website design", "Exact", "Transactional", ""),
    ("webdesign_convert", "lead generating website", "Exact", "Transactional", ""),
    ("webdesign_convert", "website to get more clients", "Exact", "Transactional", ""),
    ("webdesign_convert", "website for lead generation", "Exact", "Transactional", ""),
    ("webdesign_convert", "web design that converts", "Phrase", "Transactional", ""),
    ("webdesign_convert", "website that generates leads", "Phrase", "Transactional", ""),
    ("webdesign_convert", "conversion focused web design", "Phrase", "Transactional", ""),
    ("webdesign_convert", "high converting website", "Phrase", "Transactional", ""),
    ("webdesign_convert", "lead generation website design", "Phrase", "Transactional", ""),
]

CRO_KEYWORDS = [
    # cro_agency
    ("cro_agency", "conversion rate optimization agency", "Exact", "Transactional", "Core term"),
    ("cro_agency", "cro agency", "Exact", "Transactional", "Core term"),
    ("cro_agency", "cro company", "Exact", "Transactional", ""),
    ("cro_agency", "cro consultant", "Exact", "Transactional", ""),
    ("cro_agency", "hire cro specialist", "Exact", "Transactional", "High intent"),
    ("cro_agency", "conversion optimization agency", "Exact", "Transactional", ""),
    ("cro_agency", "cro agency canada", "Exact", "Transactional + Geo", ""),
    ("cro_agency", "website conversion agency", "Exact", "Transactional", ""),
    ("cro_agency", "conversion rate optimization agency", "Phrase", "Transactional", ""),
    ("cro_agency", "cro agency canada", "Phrase", "Transactional + Geo", ""),
    ("cro_agency", "conversion optimization agency", "Phrase", "Transactional", ""),
    ("cro_agency", "cro consultant canada", "Phrase", "Transactional + Geo", ""),
    ("cro_agency", "website cro services", "Phrase", "Commercial", ""),
    ("cro_agency", "conversion rate optimization services", "Phrase", "Commercial", ""),
    # cro_notconverting
    ("cro_notconverting", "website not converting", "Exact", "Problem/Symptom", "High intent signal"),
    ("cro_notconverting", "website not generating leads", "Exact", "Problem/Symptom", "High intent signal"),
    ("cro_notconverting", "increase website conversions", "Exact", "Transactional", ""),
    ("cro_notconverting", "improve website conversion rate", "Exact", "Transactional", ""),
    ("cro_notconverting", "how to get more leads from website", "Exact", "High Intent", ""),
    ("cro_notconverting", "website visitors not converting", "Exact", "Problem/Symptom", ""),
    ("cro_notconverting", "website not getting leads", "Exact", "Problem/Symptom", ""),
    ("cro_notconverting", "fix website conversion rate", "Exact", "Transactional", ""),
    ("cro_notconverting", "website not converting", "Phrase", "Problem/Symptom", ""),
    ("cro_notconverting", "website not generating leads", "Phrase", "Problem/Symptom", ""),
    ("cro_notconverting", "increase website conversions", "Phrase", "Transactional", ""),
    ("cro_notconverting", "improve website conversion rate", "Phrase", "Transactional", ""),
    ("cro_notconverting", "get more leads from website", "Phrase", "High Intent", ""),
    ("cro_notconverting", "website conversion help", "Phrase", "Commercial", ""),
    ("cro_notconverting", "why is my website not converting", "Phrase", "Problem/Symptom", ""),
    ("cro_notconverting", "website conversion improvement", "Phrase", "Transactional", ""),
    # cro_audit
    ("cro_audit", "conversion rate audit", "Exact", "Transactional", ""),
    ("cro_audit", "website cro audit", "Exact", "Transactional", ""),
    ("cro_audit", "conversion tracking setup", "Exact", "Transactional", ""),
    ("cro_audit", "website audit conversion", "Exact", "Commercial", ""),
    ("cro_audit", "ga4 setup agency", "Exact", "Transactional", ""),
    ("cro_audit", "google analytics setup service", "Exact", "Transactional", ""),
    ("cro_audit", "website performance audit", "Exact", "Commercial", ""),
    ("cro_audit", "conversion rate audit", "Phrase", "Transactional", ""),
    ("cro_audit", "website cro audit", "Phrase", "Transactional", ""),
    ("cro_audit", "conversion tracking setup", "Phrase", "Transactional", ""),
    ("cro_audit", "google analytics setup", "Phrase", "Transactional", ""),
    ("cro_audit", "ga4 setup service", "Phrase", "Transactional", ""),
    ("cro_audit", "website analytics setup", "Phrase", "Transactional", ""),
    ("cro_audit", "why is my website not converting", "Phrase", "Problem", ""),
]

# RSAs: (campaign, ad_group, type, position, text, char_count)
RSA_HEADLINES = [
    # === webdesign_agency ===
    ("webdesign", "webdesign_agency", "Headline", "H1 (KW)", "Web Design Agency Canada", 24),
    ("webdesign", "webdesign_agency", "Headline", "H2 (KW)", "Website Design Company", 22),
    ("webdesign", "webdesign_agency", "Headline", "H3 (KW)", "Professional Web Design", 23),
    ("webdesign", "webdesign_agency", "Headline", "H4 (Benefit)", "Websites That Generate Leads", 28),
    ("webdesign", "webdesign_agency", "Headline", "H5 (Benefit)", "Turn Clicks Into Customers", 26),
    ("webdesign", "webdesign_agency", "Headline", "H6 (Benefit)", "Design, Speed and Tracking", 26),
    ("webdesign", "webdesign_agency", "Headline", "H7 (Benefit)", "Built to Convert and Track", 25),
    ("webdesign", "webdesign_agency", "Headline", "H8 (Proof)", "11 Canadian Brands Scaled", 25),
    ("webdesign", "webdesign_agency", "Headline", "H9 (Proof) ⚠️", "Rated [X.X]/5 on Trustpilot", 27),
    ("webdesign", "webdesign_agency", "Headline", "H10 (Proof)", "Portfolio of 11 Client Wins", 27),
    ("webdesign", "webdesign_agency", "Headline", "H11 (CTA)", "Book a Free Strategy Call", 25),
    ("webdesign", "webdesign_agency", "Headline", "H12 (CTA)", "See If You Qualify Today", 24),
    ("webdesign", "webdesign_agency", "Headline", "H13 (CTA)", "Get Your Free Strategy Call", 27),
    ("webdesign", "webdesign_agency", "Headline", "H14 (Diff)", "Not Just Pretty, It Converts", 28),
    ("webdesign", "webdesign_agency", "Headline", "H15 (Diff)", "Results-Driven Web Design", 25),
    ("webdesign", "webdesign_agency", "Description", "D1 Value Prop", "Websites built to convert. Design, speed and tracking in one service. Book your free strategy call.", 83),
    ("webdesign", "webdesign_agency", "Description", "D2 Proof", "11+ Canadian brands trust Hoski for web design that converts. View our case studies.", 83),
    ("webdesign", "webdesign_agency", "Description", "D3 Diff", "We combine design, speed and conversion tracking in every project. See if you qualify.", 87),
    ("webdesign", "webdesign_agency", "Description", "D4 Objection", "Not sure if you need a redesign? We'll review your current site on a free strategy call.", 87),
    # === webdesign_redesign ===
    ("webdesign", "webdesign_redesign", "Headline", "H1 (KW)", "Website Redesign Agency", 23),
    ("webdesign", "webdesign_redesign", "Headline", "H2 (KW)", "Redesign Your Website", 21),
    ("webdesign", "webdesign_redesign", "Headline", "H3 (KW)", "Website Revamp Experts", 21),
    ("webdesign", "webdesign_redesign", "Headline", "H4 (Benefit)", "Redesign That Converts", 22),
    ("webdesign", "webdesign_redesign", "Headline", "H5 (Benefit)", "New Site, More Leads", 20),
    ("webdesign", "webdesign_redesign", "Headline", "H6 (Benefit)", "From Pretty to Profitable", 25),
    ("webdesign", "webdesign_redesign", "Headline", "H7 (Benefit)", "Redesign With ROI Tracking", 26),
    ("webdesign", "webdesign_redesign", "Headline", "H8 (Proof)", "Portfolio of 11 Redesigns", 25),
    ("webdesign", "webdesign_redesign", "Headline", "H9 (Proof) ⚠️", "Rated [X.X]/5 on Trustpilot", 27),
    ("webdesign", "webdesign_redesign", "Headline", "H10 (Proof)", "Results Across 5 Industries", 27),
    ("webdesign", "webdesign_redesign", "Headline", "H11 (CTA)", "Book a Free Strategy Call", 25),
    ("webdesign", "webdesign_redesign", "Headline", "H12 (CTA)", "Get a Free Site Review", 22),
    ("webdesign", "webdesign_redesign", "Headline", "H13 (CTA)", "See Your Redesign Potential", 27),
    ("webdesign", "webdesign_redesign", "Headline", "H14 (Diff)", "Same Budget, Better Results", 27),
    ("webdesign", "webdesign_redesign", "Headline", "H15 (Diff)", "No Generic Templates", 20),
    ("webdesign", "webdesign_redesign", "Description", "D1 Value Prop", "Redesign built to convert. Speed and tracking included. Book a free strategy call now.", 85),
    ("webdesign", "webdesign_redesign", "Description", "D2 Proof", "11+ Canadian brands redesigned with Hoski for results they can track. See our portfolio.", 87),
    ("webdesign", "webdesign_redesign", "Description", "D3 Diff", "We don't just make your site look better. We make it work better. See if you qualify.", 84),
    ("webdesign", "webdesign_redesign", "Description", "D4 Objection", "Worried your current site can't be saved? Get a free site review on your strategy call.", 87),
    # === webdesign_convert ===
    ("webdesign", "webdesign_convert", "Headline", "H1 (KW)", "Conversion Focused Web Design", 29),
    ("webdesign", "webdesign_convert", "Headline", "H2 (KW)", "Web Design That Converts", 23),
    ("webdesign", "webdesign_convert", "Headline", "H3 (KW)", "Websites That Generate Leads", 28),
    ("webdesign", "webdesign_convert", "Headline", "H4 (Benefit)", "Turn Traffic Into Clients", 24),
    ("webdesign", "webdesign_convert", "Headline", "H5 (Benefit)", "Sites That Work While You Sleep", 30),
    ("webdesign", "webdesign_convert", "Headline", "H6 (Benefit)", "Make Every Visitor Count", 23),
    ("webdesign", "webdesign_convert", "Headline", "H7 (Benefit)", "Visitors In, Customers Out", 27),
    ("webdesign", "webdesign_convert", "Headline", "H8 (Proof)", "Results for 11 Canadian Brands", 30),
    ("webdesign", "webdesign_convert", "Headline", "H9 (Proof) ⚠️", "Rated [X.X]/5 on Trustpilot", 27),
    ("webdesign", "webdesign_convert", "Headline", "H10 (CTA)", "Book a Free Strategy Call", 25),
    ("webdesign", "webdesign_convert", "Headline", "H11 (CTA)", "Get a Free Conversion Audit", 27),
    ("webdesign", "webdesign_convert", "Headline", "H12 (CTA)", "See Your Conversion Score", 24),
    ("webdesign", "webdesign_convert", "Headline", "H13 (Diff)", "Design and Tracking Combined", 28),
    ("webdesign", "webdesign_convert", "Headline", "H14 (Diff)", "Data-Backed Conversion Design", 29),
    ("webdesign", "webdesign_convert", "Headline", "H15 (Diff)", "Track Every Lead Source", 22),
    ("webdesign", "webdesign_convert", "Description", "D1 Value Prop", "Website built to convert visitors. Tracking and speed included. Book a free strategy call.", 90),
    ("webdesign", "webdesign_convert", "Description", "D2 Proof", "11+ Canadian brands turn website traffic into leads with Hoski. View case studies.", 82),
    ("webdesign", "webdesign_convert", "Description", "D3 Diff", "Most agencies design for aesthetics. We design for conversions. See the difference.", 83),
    ("webdesign", "webdesign_convert", "Description", "D4 Objection", "Not sure your site can generate leads? Get a free review on your strategy call.", 79),
    # === cro_agency ===
    ("cro", "cro_agency", "Headline", "H1 (KW)", "Conversion Rate Optimization", 28),
    ("cro", "cro_agency", "Headline", "H2 (KW)", "CRO Agency Canada", 18),
    ("cro", "cro_agency", "Headline", "H3 (KW)", "Hire a CRO Specialist", 21),
    ("cro", "cro_agency", "Headline", "H4 (Benefit)", "Turn Your Traffic Into Leads", 28),
    ("cro", "cro_agency", "Headline", "H5 (Benefit)", "Stop Losing Leads to Poor UX", 28),
    ("cro", "cro_agency", "Headline", "H6 (Benefit)", "More Conversions Same Traffic", 29),
    ("cro", "cro_agency", "Headline", "H7 (Benefit)", "Why Visitors Don't Convert", 25),
    ("cro", "cro_agency", "Headline", "H8 (Proof)", "11 Canadian Brands Optimized", 28),
    ("cro", "cro_agency", "Headline", "H9 (Proof) ⚠️", "Rated [X.X]/5 on Trustpilot", 27),
    ("cro", "cro_agency", "Headline", "H10 (CTA)", "Book a Free Strategy Call", 25),
    ("cro", "cro_agency", "Headline", "H11 (CTA)", "Get a Free CRO Audit", 21),
    ("cro", "cro_agency", "Headline", "H12 (CTA)", "See Your Conversion Score", 24),
    ("cro", "cro_agency", "Headline", "H13 (Diff)", "CRO and Web Design Combined", 26),
    ("cro", "cro_agency", "Headline", "H14 (Diff)", "Design, Speed and Tracking", 26),
    ("cro", "cro_agency", "Headline", "H15 (Diff)", "Track Every Lead Source", 22),
    ("cro", "cro_agency", "Description", "D1 Value Prop", "We audit, redesign and track conversions in one service. Book a free strategy call.", 83),
    ("cro", "cro_agency", "Description", "D2 Proof", "11+ Canadian brands improved their conversion rates with Hoski. View our case studies.", 87),
    ("cro", "cro_agency", "Description", "D3 Diff", "We focus on converting your existing traffic, not just driving more of it. Book a call.", 87),
    ("cro", "cro_agency", "Description", "D4 Objection", "Getting traffic but not enough leads? Get a free conversion review on your strategy call.", 89),
    # === cro_notconverting ===
    ("cro", "cro_notconverting", "Headline", "H1 (KW)", "Website Not Converting?", 23),
    ("cro", "cro_notconverting", "Headline", "H2 (KW)", "Improve Your Conversion Rate", 28),
    ("cro", "cro_notconverting", "Headline", "H3 (KW)", "Get More Leads From Your Site", 29),
    ("cro", "cro_notconverting", "Headline", "H4 (Benefit)", "Traffic Without Leads Fails", 27),
    ("cro", "cro_notconverting", "Headline", "H5 (Benefit)", "Fix What's Killing Your Leads", 29),
    ("cro", "cro_notconverting", "Headline", "H6 (Benefit)", "Your Traffic Deserves Better", 28),
    ("cro", "cro_notconverting", "Headline", "H7 (Benefit)", "Turn Visitors Into Clients", 26),
    ("cro", "cro_notconverting", "Headline", "H8 (Proof)", "11 Brands Fixed With Hoski", 25),
    ("cro", "cro_notconverting", "Headline", "H9 (Proof) ⚠️", "Rated [X.X]/5 on Trustpilot", 27),
    ("cro", "cro_notconverting", "Headline", "H10 (CTA)", "Book a Free Strategy Call", 25),
    ("cro", "cro_notconverting", "Headline", "H11 (CTA)", "Get a Free Conversion Review", 28),
    ("cro", "cro_notconverting", "Headline", "H12 (CTA)", "Find Out Why You Lose Leads", 26),
    ("cro", "cro_notconverting", "Headline", "H13 (Diff)", "Audit, Redesign and Track", 25),
    ("cro", "cro_notconverting", "Headline", "H14 (Diff)", "CRO Plus Web Design, Combined", 29),
    ("cro", "cro_notconverting", "Headline", "H15 (Diff)", "No More Guessing Why", 21),
    ("cro", "cro_notconverting", "Description", "D1 Value Prop", "Site not converting? We audit, redesign and track to fix it. Book a free strategy call.", 87),
    ("cro", "cro_notconverting", "Description", "D2 Proof", "11+ Canadian brands improved their website conversions with Hoski. View our case studies.", 89),
    ("cro", "cro_notconverting", "Description", "D3 Diff", "We combine CRO analysis with web design and tracking, so you know exactly what to fix.", 86),
    ("cro", "cro_notconverting", "Description", "D4 Objection", "Your visitors are telling you something. Let us show you what. Free conversion review.", 85),
    # === cro_audit ===
    ("cro", "cro_audit", "Headline", "H1 (KW)", "Conversion Rate Audit", 22),
    ("cro", "cro_audit", "Headline", "H2 (KW)", "Website CRO Audit", 17),
    ("cro", "cro_audit", "Headline", "H3 (KW)", "Conversion Tracking Setup", 25),
    ("cro", "cro_audit", "Headline", "H4 (Benefit)", "Know Where Leads Come From", 26),
    ("cro", "cro_audit", "Headline", "H5 (Benefit)", "See Exactly What's Not Working", 30),
    ("cro", "cro_audit", "Headline", "H6 (Benefit)", "Track Every Lead and Sale", 25),
    ("cro", "cro_audit", "Headline", "H7 (Benefit)", "Fix Your Conversion Funnel", 26),
    ("cro", "cro_audit", "Headline", "H8 (Proof)", "11 Brands Audited and Fixed", 27),
    ("cro", "cro_audit", "Headline", "H9 (Proof) ⚠️", "Rated [X.X]/5 on Trustpilot", 27),
    ("cro", "cro_audit", "Headline", "H10 (CTA)", "Book a Free Strategy Call", 25),
    ("cro", "cro_audit", "Headline", "H11 (CTA)", "Get Your Free CRO Audit", 23),
    ("cro", "cro_audit", "Headline", "H12 (CTA)", "Start With a Free Site Review", 29),
    ("cro", "cro_audit", "Headline", "H13 (Diff)", "Audit, Redesign and Track", 25),
    ("cro", "cro_audit", "Headline", "H14 (Diff)", "Analytics Setup Included", 24),
    ("cro", "cro_audit", "Headline", "H15 (Diff)", "Know Your Conversion Rate", 24),
    ("cro", "cro_audit", "Description", "D1 Value Prop", "Site audit, conversion fixes and tracking setup in one service. Book a free strategy call.", 90),
    ("cro", "cro_audit", "Description", "D2 Proof", "11+ Canadian brands use Hoski for conversion audits. View our case studies today.", 81),
    ("cro", "cro_audit", "Description", "D3 Diff", "We set up GA4, conversion tracking and heatmaps, then redesign based on real data.", 82),
    ("cro", "cro_audit", "Description", "D4 Objection", "Not sure what's hurting your conversion rate? Get a full audit on your free strategy call.", 89),
]

SITELINKS = [
    # (campaign, text, final_url, desc1, desc2)
    ("webdesign", "See Client Case Studies", "/case-studies", "11 Canadian brands scaled", "E-commerce, dental, aesthetics and more"),
    ("webdesign", "What's Included", "/web-desk", "Design, speed and tracking", "All in one service"),
    ("webdesign", "How Our Process Works", "/web-desk", "3-step process explained", "Audit, build and track"),
    ("webdesign", "Book a Strategy Call", "/web-desk#book", "30 minutes, no commitment", "Review your site free"),
    ("webdesign", "Web Desk Service", "/web-desk", "Conversion-focused web design", "Built for growing brands"),
    ("webdesign", "About Hoski", "/about", "Canadian-based agency", "Trusted by local and luxury brands"),
    ("cro", "See Client Case Studies", "/case-studies", "11 Canadian brands scaled", "Before and after results"),
    ("cro", "Get a Free CRO Review", "/web-desk#book", "We'll audit your site live", "On a 30-min strategy call"),
    ("cro", "How Our CRO Process Works", "/web-desk", "Audit, redesign and track", "Three steps to more leads"),
    ("cro", "Book a Strategy Call", "/web-desk#book", "30 minutes, no commitment", "Walk away with a clear plan"),
    ("cro", "Web Desk Service", "/web-desk", "CRO and web design combined", "Speed + design + tracking"),
    ("cro", "About Hoski", "/about", "Canadian-based CRO agency", "Results-first, not design-first"),
]

CALLOUTS = [
    ("webdesign", "Conversion Tracking Included"),
    ("webdesign", "Speed Optimization Included"),
    ("webdesign", "No Long-Term Contracts"),
    ("webdesign", "Canadian-Based Agency"),
    ("webdesign", "Trusted by 11+ Brands"),
    ("webdesign", "Results-First Approach"),
    ("webdesign", "Free Strategy Call"),
    ("webdesign", "No DIY Templates"),
    ("webdesign", "Design + Speed + Tracking"),
    ("webdesign", "Rated on Trustpilot"),
    ("cro", "Free Conversion Audit on Call"),
    ("cro", "Analytics Setup Included"),
    ("cro", "GA4 and Tracking Configured"),
    ("cro", "No Long-Term Contracts"),
    ("cro", "Canadian-Based Agency"),
    ("cro", "Trusted by 11+ Brands"),
    ("cro", "CRO Plus Web Design in One"),
    ("cro", "No More Guessing Why It Fails"),
    ("cro", "Data-Backed Recommendations"),
    ("cro", "Rated on Trustpilot"),
]

NEGATIVES_ACCOUNT = [
    ("Account", "Broad", "jobs", "Employment"),
    ("Account", "Broad", "careers", "Employment"),
    ("Account", "Broad", "hiring", "Employment"),
    ("Account", "Broad", "salary", "Employment"),
    ("Account", "Broad", "resume", "Employment"),
    ("Account", "Broad", "how to", "Informational/DIY"),
    ("Account", "Broad", "tutorial", "Informational/DIY"),
    ("Account", "Broad", "free", "DIY"),
    ("Account", "Broad", "diy", "DIY"),
    ("Account", "Broad", "do it yourself", "DIY"),
    ("Account", "Broad", "course", "Educational"),
    ("Account", "Broad", "training", "Educational"),
    ("Account", "Broad", "school", "Educational"),
    ("Account", "Broad", "degree", "Educational"),
    ("Account", "Broad", "certification", "Educational"),
    ("Account", "Broad", "internship", "Employment"),
    ("Account", "Broad", "learn", "Educational"),
    ("Account", "Broad", "template", "DIY"),
    ("Account", "Broad", "wix", "DIY platform"),
    ("Account", "Broad", "squarespace", "DIY platform"),
    ("Account", "Broad", "weebly", "DIY platform"),
    ("Account", "Broad", "elementor", "DIY platform"),
    ("Account", "Broad", "cheap", "Price-driven / low ATV"),
    ("Account", "Broad", "affordable", "Price-driven / low ATV"),
    ("Account", "Broad", "low cost", "Price-driven / low ATV"),
    ("Account", "Broad", "student", "Not target market"),
    ("Account", "Broad", "nonprofit", "Not target market"),
    ("Account", "Broad", "hobby", "Not target market"),
    ("Account", "Phrase", "website builder", "DIY intent"),
    ("Account", "Phrase", "build your own website", "DIY intent"),
    ("Account", "Phrase", "free website", "DIY intent"),
    ("Account", "Phrase", "web design school", "Educational"),
    ("Account", "Phrase", "web design jobs", "Employment"),
    ("Account", "Phrase", "web design course", "Educational"),
    ("Account", "Phrase", "web design tutorial", "DIY"),
    ("Account", "Phrase", "shopify theme", "DIY/product — not service"),
    ("Account", "Phrase", "wordpress theme", "DIY/product"),
    ("Account", "Phrase", "make my own website", "DIY intent"),
]

NEGATIVES_CAMPAIGN = [
    ("hoski_webdesign_search_1apr26", "Campaign", "Phrase", "thrillx", "Competitor"),
    ("hoski_webdesign_search_1apr26", "Campaign", "Phrase", "parachute design", "Competitor"),
    ("hoski_webdesign_search_1apr26", "Campaign", "Phrase", "consultus", "Competitor"),
    ("hoski_webdesign_search_1apr26", "Campaign", "Phrase", "nvision", "Competitor"),
    ("hoski_webdesign_search_1apr26", "Campaign", "Phrase", "mediaforce", "Competitor"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "thrillx", "Competitor"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "parachute design", "Competitor"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "consultus", "Competitor"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "hotjar", "Software — not service"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "cro software", "Software — not service"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "cro plugin", "Software — not service"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "unbounce", "Software — not service"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "instapage", "Software — not service"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "cro course", "Educational"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "cro certification", "Educational"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "freelancer", "Wrong buyer type"),
    ("hoski_cro_search_1apr26", "Campaign", "Phrase", "fiverr", "Wrong buyer type"),
]

NEGATIVES_ADGROUP = [
    ("hoski_webdesign_search_1apr26", "webdesign_agency", "Exact", "redesign", "Route to webdesign_redesign"),
    ("hoski_webdesign_search_1apr26", "webdesign_agency", "Exact", "converts", "Route to webdesign_convert"),
    ("hoski_webdesign_search_1apr26", "webdesign_agency", "Exact", "conversion", "Route to webdesign_convert"),
    ("hoski_webdesign_search_1apr26", "webdesign_redesign", "Exact", "converts", "Route to webdesign_convert"),
    ("hoski_webdesign_search_1apr26", "webdesign_redesign", "Exact", "conversion", "Route to webdesign_convert"),
    ("hoski_webdesign_search_1apr26", "webdesign_convert", "Exact", "redesign", "Route to webdesign_redesign"),
    ("hoski_webdesign_search_1apr26", "webdesign_convert", "Exact", "agency", "Route to webdesign_agency"),
    ("hoski_cro_search_1apr26", "cro_agency", "Exact", "not converting", "Route to cro_notconverting"),
    ("hoski_cro_search_1apr26", "cro_agency", "Exact", "audit", "Route to cro_audit"),
    ("hoski_cro_search_1apr26", "cro_agency", "Exact", "tracking", "Route to cro_audit"),
    ("hoski_cro_search_1apr26", "cro_notconverting", "Exact", "agency", "Route to cro_agency"),
    ("hoski_cro_search_1apr26", "cro_notconverting", "Exact", "audit", "Route to cro_audit"),
    ("hoski_cro_search_1apr26", "cro_audit", "Exact", "agency", "Route to cro_agency"),
    ("hoski_cro_search_1apr26", "cro_audit", "Exact", "not converting", "Route to cro_notconverting"),
]

CHECKLIST = [
    ("BLOCKER", "Conversion tracking confirmed firing (strategy call booked event)"),
    ("BLOCKER", "CallRail phone number added to call assets (ACC019cc4e94fca7082bca220109c69d5ef)"),
    ("BLOCKER", "Trustpilot score confirmed — update all H9 rows marked ⚠️ with actual rating"),
    ("HIGH", "Account-level universal negative list uploaded"),
    ("HIGH", "Campaign-level negatives uploaded to each campaign"),
    ("HIGH", "Ad group cross-negatives applied"),
    ("HIGH", "Display Network checkbox confirmed OFF on both campaigns"),
    ("HIGH", "Minimum 2 RSA variants per ad group built (12 RSAs total across both campaigns — this sheet has 1 per group, build a second variant)"),
    ("HIGH", "Sitelinks live with descriptions — minimum 4 per campaign"),
    ("HIGH", "Callout extensions added — minimum 8 per campaign"),
    ("HIGH", "Structured snippets added"),
    ("HIGH", "URL tracking parameters added to final URLs"),
    ("HIGH", "Geo bid adjustments set (+15% Montreal, Toronto, Vancouver, Calgary)"),
    ("MEDIUM", "Ad rotation set to Optimize"),
    ("MEDIUM", "Search Partners ON"),
    ("MEDIUM", "Language set to English + French"),
    ("MEDIUM", "Billing verified in Google Ads account"),
    ("POST-LAUNCH", "Week 1: Pull search term report, add irrelevant queries as phrase negatives"),
    ("POST-LAUNCH", "Week 2: Check impression share per ad group — if <50%, assess budget/keyword tightness"),
    ("POST-LAUNCH", "Week 4: Pause RSA assets rated Poor after 2,000+ impressions"),
    ("POST-LAUNCH", "Day 30: Review geo performance — increase bid adjustments if city outperforms"),
    ("POST-LAUNCH", "After 30+ conversions per campaign: switch bid strategy to Maximize Conversions"),
]

# ── Workbook construction ─────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

def add_sheet(title, tab_color=C_DARK):
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = tab_color
    return ws

# ── Sheet 1: Campaign Settings ────────────────────────────────────────────────
ws = add_sheet("01 Campaign Settings", "1A1A2E")
freeze(ws)
ws.row_dimensions[1].height = 30
write_header_row(ws, 1,
    ["Setting", "hoski_webdesign_search_1apr26", "hoski_cro_search_1apr26"],
    bg=C_DARK, fg=C_WHITE, size=11)
for i, row in enumerate(CAMPAIGN_SETTINGS[1:], start=2):
    bg = C_LIGHT_BLUE if i % 2 == 0 else C_WHITE
    write_data_row(ws, i, row, bg=bg)
set_col_widths(ws, [36, 55, 55])
ws.row_dimensions[1].height = 30

# ── Sheet 2: WD Keywords ─────────────────────────────────────────────────────
ws = add_sheet("02 WD Keywords", "0F3460")
freeze(ws)
merge_header(ws, 1, "hoski_webdesign_search_1apr26 — Keywords", 1, 5, bg=C_DARK)
write_header_row(ws, 2, ["Ad Group", "Keyword", "Match Type", "Intent", "Notes"],
                 bg=C_ACCENT)
ad_group_colors = {
    "webdesign_agency": C_LIGHT_BLUE,
    "webdesign_redesign": "E8EAF6",
    "webdesign_convert": "E8F5E9",
}
for i, row in enumerate(WD_KEYWORDS, start=3):
    bg = ad_group_colors.get(row[0], C_WHITE)
    match_color = "E3F2FD" if row[2] == "Exact" else "FFF9C4"
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        c.fill = fill(bg)
        c.font = font()
        c.alignment = left()
        c.border = border_thin()
    ws.cell(row=i, column=3).fill = fill(match_color)
set_col_widths(ws, [22, 38, 14, 22, 28])

# ── Sheet 3: CRO Keywords ─────────────────────────────────────────────────────
ws = add_sheet("03 CRO Keywords", "0F3460")
freeze(ws)
merge_header(ws, 1, "hoski_cro_search_1apr26 — Keywords", 1, 5, bg=C_DARK)
write_header_row(ws, 2, ["Ad Group", "Keyword", "Match Type", "Intent", "Notes"],
                 bg=C_ACCENT)
ad_group_colors2 = {
    "cro_agency": C_LIGHT_BLUE,
    "cro_notconverting": "E8EAF6",
    "cro_audit": "E8F5E9",
}
for i, row in enumerate(CRO_KEYWORDS, start=3):
    bg = ad_group_colors2.get(row[0], C_WHITE)
    match_color = "E3F2FD" if row[2] == "Exact" else "FFF9C4"
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        c.fill = fill(bg)
        c.font = font()
        c.alignment = left()
        c.border = border_thin()
    ws.cell(row=i, column=3).fill = fill(match_color)
set_col_widths(ws, [22, 42, 14, 22, 28])

# ── Sheet 4: RSA Ads ─────────────────────────────────────────────────────────
ws = add_sheet("04 RSA Ads", "E94560")
freeze(ws)
merge_header(ws, 1, "RSA Assets — All Campaigns and Ad Groups", 1, 6, bg=C_DARK)
write_header_row(ws, 2,
    ["Campaign", "Ad Group", "Type", "Position / Category", "Text", "Char Count"],
    bg=C_ACCENT)

hl_color = "E3F2FD"
desc_color = "FFF9C4"
warn_color = "FFEBEE"
ad_group_fill = {
    "webdesign_agency": "EDE7F6",
    "webdesign_redesign": "E8EAF6",
    "webdesign_convert": "E8F5E9",
    "cro_agency": "E1F5FE",
    "cro_notconverting": "FFF8E1",
    "cro_audit": "F3E5F5",
}

for i, (camp, ag, atype, pos, text, chars) in enumerate(RSA_HEADLINES, start=3):
    row_bg = ad_group_fill.get(ag, C_WHITE)
    if atype == "Description":
        row_bg = desc_color
    is_warn = "⚠️" in pos

    for col, val in enumerate([camp, ag, atype, pos, text, chars], start=1):
        c = ws.cell(row=i, column=col, value=val)
        c.fill = fill(warn_color if is_warn else row_bg)
        c.font = font(bold=(col == 5), color="C62828" if is_warn else C_DARK)
        c.alignment = left()
        c.border = border_thin()

    # char count — colour-code over limit
    char_cell = ws.cell(row=i, column=6)
    limit = 30 if atype == "Headline" else 90
    if chars > limit:
        char_cell.fill = fill("FFEBEE")
        char_cell.font = font(bold=True, color="C62828")
    elif chars >= limit - 3:
        char_cell.fill = fill("FFF9C4")
        char_cell.font = font(bold=True, color="F57F17")
    else:
        char_cell.fill = fill("E8F5E9")
        char_cell.font = font(color="1B5E20")

set_col_widths(ws, [12, 22, 14, 20, 68, 12])
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 22

# ── Sheet 5: Assets ──────────────────────────────────────────────────────────
ws = add_sheet("05 Assets", "16213E")
freeze(ws, "A3")

# Sitelinks
merge_header(ws, 1, "SITELINKS", 1, 5, bg=C_DARK)
write_header_row(ws, 2,
    ["Campaign", "Sitelink Text", "Final URL", "Description Line 1", "Description Line 2"],
    bg=C_ACCENT)
for i, row in enumerate(SITELINKS, start=3):
    bg = C_LIGHT_BLUE if row[0] == "webdesign" else "E8EAF6"
    write_data_row(ws, i, row, bg=bg)

sl_end = 3 + len(SITELINKS)

# Callouts
merge_header(ws, sl_end + 1, "CALLOUT EXTENSIONS", 1, 5, bg=C_MID)
write_header_row(ws, sl_end + 2, ["Campaign", "Callout Text", "", "", ""],
                 bg=C_ACCENT)
for i, (camp, text) in enumerate(CALLOUTS, start=sl_end + 3):
    bg = C_LIGHT_BLUE if camp == "webdesign" else "E8EAF6"
    write_data_row(ws, i, [camp, text, "", "", ""], bg=bg)

co_end = sl_end + 3 + len(CALLOUTS)

# Structured Snippets
snip_row = co_end + 1
merge_header(ws, snip_row, "STRUCTURED SNIPPETS", 1, 5, bg=C_MID)
write_header_row(ws, snip_row + 1, ["Campaign", "Header Type", "Value 1", "Value 2", "Value 3"], bg=C_ACCENT)
snippets = [
    ("webdesign", "Services", "Web Design", "Landing Page Design", "CRO Audit"),
    ("webdesign", "Services (cont.)", "Speed Optimization", "Conversion Tracking", ""),
    ("webdesign", "Industries", "E-Commerce", "Dental", "Aesthetics"),
    ("webdesign", "Industries (cont.)", "Automotive", "Home Improvement", ""),
    ("cro", "Services", "CRO Audit", "Conversion Tracking", "GA4 Setup"),
    ("cro", "Services (cont.)", "Website Redesign", "Speed Optimization", ""),
    ("cro", "Industries", "E-Commerce", "Dental", "Aesthetics"),
    ("cro", "Types", "CRO Audit", "Heatmap Analysis", "Analytics Setup"),
]
for i, row in enumerate(snippets, start=snip_row + 2):
    bg = C_LIGHT_BLUE if row[0] == "webdesign" else "E8EAF6"
    write_data_row(ws, i, row, bg=bg)

set_col_widths(ws, [14, 30, 35, 35, 35])

# ── Sheet 6: Negatives ────────────────────────────────────────────────────────
ws = add_sheet("06 Negatives", "C62828")
freeze(ws, "A3")

# Account negatives
merge_header(ws, 1, "ACCOUNT-LEVEL NEGATIVES — List Name: hoski_universal_negatives", 1, 4, bg="B71C1C", fg=C_WHITE)
write_header_row(ws, 2, ["Level", "Match Type", "Keyword", "Reason"], bg="C62828", fg=C_WHITE)
for i, row in enumerate(NEGATIVES_ACCOUNT, start=3):
    bg = "FFEBEE" if i % 2 == 0 else "FFF5F5"
    write_data_row(ws, i, row, bg=bg)

acc_end = 3 + len(NEGATIVES_ACCOUNT)

# Campaign negatives
merge_header(ws, acc_end + 1, "CAMPAIGN-LEVEL NEGATIVES", 1, 4, bg="B71C1C", fg=C_WHITE)
write_header_row(ws, acc_end + 2, ["Campaign", "Level", "Match Type", "Keyword / Reason"], bg="C62828", fg=C_WHITE)
for i, row in enumerate(NEGATIVES_CAMPAIGN, start=acc_end + 3):
    bg = "FFEBEE" if i % 2 == 0 else "FFF5F5"
    write_data_row(ws, i, [row[0], row[1], row[2], f"{row[3]} — {row[4]}"], bg=bg)

camp_end = acc_end + 3 + len(NEGATIVES_CAMPAIGN)

# Ad group cross-negatives
merge_header(ws, camp_end + 1, "AD GROUP CROSS-NEGATIVES", 1, 4, bg="B71C1C", fg=C_WHITE)
write_header_row(ws, camp_end + 2, ["Campaign", "Ad Group", "Match Type", "Keyword — Reason"], bg="C62828", fg=C_WHITE)
for i, row in enumerate(NEGATIVES_ADGROUP, start=camp_end + 3):
    bg = "FFEBEE" if i % 2 == 0 else "FFF5F5"
    write_data_row(ws, i, [row[0], row[1], row[2], f"{row[3]} — {row[4]}"], bg=bg)

set_col_widths(ws, [40, 22, 14, 45])

# ── Sheet 7: Pre-Launch Checklist ─────────────────────────────────────────────
ws = add_sheet("07 Pre-Launch Checklist", "1B5E20")
freeze(ws)
merge_header(ws, 1, "PRE-LAUNCH CHECKLIST — Complete before first campaign goes live", 1, 3,
             bg="1B5E20", fg=C_WHITE, size=12)
write_header_row(ws, 2, ["Priority", "Action Item", "Status"],
                 bg="2E7D32", fg=C_WHITE)

priority_colors = {
    "BLOCKER": ("B71C1C", "FFEBEE"),
    "HIGH": ("E65100", "FFF3E0"),
    "MEDIUM": ("1565C0", "E3F2FD"),
    "POST-LAUNCH": ("4A148C", "F3E5F5"),
}

for i, (priority, item) in enumerate(CHECKLIST, start=3):
    text_color, bg_color = priority_colors.get(priority, (C_DARK, C_WHITE))
    c_pri = ws.cell(row=i, column=1, value=priority)
    c_pri.fill = fill(bg_color)
    c_pri.font = font(bold=True, color=text_color)
    c_pri.alignment = center()
    c_pri.border = border_thin()

    c_item = ws.cell(row=i, column=2, value=item)
    c_item.fill = fill(bg_color)
    c_item.font = font(color=C_DARK)
    c_item.alignment = left()
    c_item.border = border_thin()

    c_status = ws.cell(row=i, column=3, value="")
    c_status.fill = fill("FAFAFA")
    c_status.border = border_thin()
    c_status.alignment = center()

set_col_widths(ws, [16, 80, 20])

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
print(f"  WD Keywords: {len(WD_KEYWORDS)} rows")
print(f"  CRO Keywords: {len(CRO_KEYWORDS)} rows")
print(f"  RSA Assets: {len(RSA_HEADLINES)} rows ({len(RSA_HEADLINES)//19} ad groups)")
print(f"  Sitelinks: {len(SITELINKS)} rows")
print(f"  Callouts: {len(CALLOUTS)} rows")
print(f"  Negatives: {len(NEGATIVES_ACCOUNT)+len(NEGATIVES_CAMPAIGN)+len(NEGATIVES_ADGROUP)} total rows")
print(f"  Checklist items: {len(CHECKLIST)}")
