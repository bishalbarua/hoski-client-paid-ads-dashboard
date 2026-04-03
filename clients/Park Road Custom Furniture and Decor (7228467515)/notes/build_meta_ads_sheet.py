"""
Build Park Road Meta Ads Campaign Setup — Excel
Generates: meta-ads-setup.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/bishalbarua/Bishal/AI/antigravity/Google Ads Manager/clients/Park Road Custom Furniture and Decor (7228467515)/notes/meta-ads-setup.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
C_DARK       = "1A237E"   # deep navy header
C_BLUE_HEAD  = "1565C0"   # Custom Furniture (C1)
C_BLUE_LIGHT = "E3F2FD"
C_GREEN_HEAD = "2E7D32"   # Cottage Furniture (C2)
C_GREEN_LIGHT= "E8F5E9"
C_AMBER_HEAD = "E65100"   # Shared / assets
C_AMBER_LIGHT= "FFF3E0"
C_PURPLE_HEAD= "6A1B9A"   # Audiences
C_PURPLE_LIGHT="F3E5F5"
C_TEAL_HEAD  = "00695C"   # Budget / Launch
C_TEAL_LIGHT = "E0F2F1"
C_WHITE      = "FFFFFF"

thin = Side(style="thin", color="BDBDBD")
def border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def hfont(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def wrap(ws, row, col, value, fill=None, bold=False, font_color="000000", size=10, h_align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = hfont(bold=bold, color=font_color, size=size)
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal=h_align)
    cell.border = border()
    if fill:
        cell.fill = hfill(fill)
    return cell

def header_row(ws, row_num, values, bg, font_color="FFFFFF", size=10):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.font = hfont(bold=True, color=font_color, size=size)
        cell.fill = hfill(bg)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border()

def section_title(ws, row, text, merge_to, color):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = hfont(bold=True, color="FFFFFF", size=11)
    cell.fill = hfill(color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ════════════════════════════════════════════════════════════════════════════
# DATA
# ════════════════════════════════════════════════════════════════════════════

# ── Campaign Settings ────────────────────────────────────────────────────────
CAMP_SETTINGS = [
    ("Campaign name",       "hoski_custom_furniture_meta_1apr26",           "hoski_cottage_furniture_meta_1apr26"),
    ("Objective",           "Leads",                                         "Leads"),
    ("Conversion event",    "CF - Lead (Meta)",                             "Cottage - Lead (Meta)"),
    ("Buying type",         "Auction",                                       "Auction"),
    ("Campaign budget",     "OFF — set at ad set level",                    "OFF — set at ad set level"),
    ("Special ad category", "None",                                          "None"),
    ("Primary landing page","go.parkroadfurniture.com/custom-furniture",    "go.parkroadfurniture.com/cottage-furniture"),
    ("Geo",                 "London ON + 60 km radius (SW Ontario)",        "Ontario — province-wide"),
    ("Total daily budget",  "$35/day ($20 + $15 across 2 ad sets)",        "$45/day ($25 + $20 across 2 ad sets)"),
    ("Monthly estimate",    "~$1,050/month",                                "~$1,350/month"),
]

# ── Ad Sets ──────────────────────────────────────────────────────────────────
ADSETS = [
    # Campaign | Ad Set Name | Audience Type | Geo | Budget | Age | Interests/Behaviors | Placements
    ("Custom Furniture",
     "custom_furniture_advantageplus_london",
     "Advantage+ (Meta optimizes)",
     "London ON + 60 km",
     "$20/day",
     "30–65+",
     "Audience controls (suggestions): Home furnishings, Custom furniture, Interior design, Homeowner",
     "Advantage+ Placements (all)"),

    ("Custom Furniture",
     "custom_furniture_interest_london",
     "Interest-based (manual)",
     "London ON, Kitchener ON, Waterloo ON, Windsor ON, Guelph ON, Woodstock ON, Sarnia ON, St. Thomas ON",
     "$15/day",
     "30–65+",
     "Interests: Home furnishings, Interior design, Home improvement, Custom furniture, Dining room furniture, Bedroom furniture, Solid wood furniture, Renovation\nBehaviors: Homeowners, Likely to move\nExclusions: Recent lead submitters (60d)",
     "Facebook Feed, Instagram Feed, FB Stories, IG Stories, IG Reels"),

    ("Cottage Furniture",
     "cottage_furniture_advantageplus_ontario",
     "Advantage+ (Meta optimizes)",
     "Ontario — province-wide",
     "$25/day",
     "35–65+",
     "Audience controls (suggestions): Cottage, Muskoka, Vacation homes, Farmhouse decor, Interior design, Homeowner",
     "Advantage+ Placements (all)"),

    ("Cottage Furniture",
     "cottage_furniture_interest_ontario",
     "Interest-based (manual)",
     "Ontario — province-wide (can exclude London ON to avoid overlap with C1)",
     "$20/day",
     "35–65+",
     "Interests: Cottage, Muskoka region, Cottage life, Vacation home, Farmhouse decor, Rustic home decor, Interior design, Dining room furniture, Home furnishings\nBehaviors: Homeowners, Engaged shoppers\nExclusions: Recent lead submitters (60d)",
     "Facebook Feed, Instagram Feed, FB Stories, IG Stories, IG Reels"),
]

# ── Custom Furniture Creatives ────────────────────────────────────────────────
CF_CREATIVES = [
    # (File Name, Visual Direction, Primary Text, Headline, Description, CTA, URL)
    (
        "Static1_Custom Furniture $750 Offer_Statics_Apr2026",
        "Wide dining table or bedroom set in a clean, well-lit London home. Warm natural light, real grain and finish visible. No clutter. No text overlay. Formats: 1:1, 4:5, 9:16.",
        "Most furniture is built for a showroom floor. Not your room.\n\nAt Park Road, every piece is made to order in London, Ontario — the size you need, the wood you want, the finish that fits your home.\n\n$750 OFF your first custom order of $5,000 or more.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "Built for Your Room. Built in London.",
        "Free in-home consultation available.",
        "Get Quote",
        "https://go.parkroadfurniture.com/custom-furniture",
    ),
    (
        "Static2_Custom Furniture $750 Offer_Statics_Apr2026",
        "Close-up on wood grain — table surface, leg joinery, or finish detail. Texture and craftsmanship are the story. Warm amber/honey tones. Formats: 1:1, 4:5.",
        "You've been in furniture stores. You've measured. Nothing is quite right.\n\nWrong size. Wrong wood. Not made to last.\n\nPark Road builds every piece to order using solid Canadian hardwood — custom size, custom finish, hand-built in our London workshop.\n\n$750 OFF orders of $5,000+. Book a free in-home consultation.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "Custom Furniture That Actually Fits",
        "Dining tables, beds, sofas — all built to order.",
        "Learn More",
        "https://go.parkroadfurniture.com/custom-furniture",
    ),
    (
        "Static3_Custom Furniture $750 Offer_Statics_Apr2026",
        "Offer-forward. '$750 OFF' prominent. Background: dark walnut or rich wood grain texture. White or cream text overlay. Formats: 1:1, 9:16 (Stories version with bold offer text centered).",
        "Your first custom order, $750 off.\n\nSolid Canadian hardwood, built to your exact size and finish in London, Ontario. Dining tables, sofas, bedroom sets, and more.\n\nClaim your discount — book a free consultation at the link below.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "$750 OFF — Custom Furniture London ON",
        "On first custom order of $5,000 or more.",
        "Claim Offer",
        "https://go.parkroadfurniture.com/custom-furniture",
    ),
    (
        "Static4_Custom Furniture $750 Offer_Statics_Apr2026",
        "Lifestyle image of a customer room with a Park Road piece — candid, lived-in, not staged. Alternatively: warm real-home dining scene. Formats: 1:1, 4:5.",
        "Big-box furniture is built to sell. Park Road furniture is built to last.\n\nEvery piece — dining tables, beds, sofas, entertainment units — is made to order in solid Canadian hardwood. No particle board. No veneers. No compromises.\n\nCustom size. Custom finish. Built in London, Ontario.\n\n$750 OFF your first order of $5,000+. Free in-home consultation included.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "Not Big-Box. Custom Built in London ON.",
        "Real solid wood, made to order.",
        "Get Quote",
        "https://go.parkroadfurniture.com/custom-furniture",
    ),
]

# ── Cottage Furniture Creatives ───────────────────────────────────────────────
C2_CREATIVES = [
    (
        "Static1_Cottage Furniture $750 Offer_Statics_March2026",
        "Warm cottage dining or living room. Natural light from a window. Wood-toned furniture — farmhouse table or solid wood sofa. Lived-in feel, not staged. Formats: 1:1, 4:5, 9:16.\nReference: motionapp.com/…/inspo/ad/966378142501506",
        "Your home can feel like a cottage retreat.\n\nCalm. Warm. Made for the way you actually live.\n\nPark Road builds custom solid wood furniture to order — any size, any finish, delivered across Ontario.\n\n$750 OFF orders of $5,000 or more. Book a free consultation.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "Bring The Cottage Feeling Home",
        "Custom solid wood furniture. Built to order.",
        "Get Quote",
        "https://go.parkroadfurniture.com/cottage-furniture",
        "Overlay: Headline — 'Bring The Cottage Feeling Home' | Body — 'A space that feels calm, inviting, and lived in' | Offer badge — '$750 OFF Orders $5,000+' | Footer — address + phone",
    ),
    (
        "Static2_Cottage Furniture $750 Offer_Statics_March2026",
        "Minimal, airy cottage living room or bedroom. Light wood tones, natural linen textures, soft morning light. No clutter. The space feels like a deep exhale. Formats: 1:1, 4:5.\nReference: tryatria.com/ad/m1308632981290620",
        "Designed to feel like a quiet escape from everything else.\n\nClean lines. Natural wood. A space that slows you down the moment you walk in.\n\nPark Road custom furniture is built to order — your size, your finish, your cottage.\n\nGet $750 OFF when you spend $5,000+.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "Clean, Calm, and Inviting",
        "Custom cottage furniture. Built to order in Ontario.",
        "Learn More",
        "https://go.parkroadfurniture.com/cottage-furniture",
        "Overlay: Headline — 'Clean, calm, and inviting' | Subhead — 'Designed to feel like a quiet escape from everything else' | Offer — 'Get $750 OFF When You Spend $5,000+' | Footer",
    ),
    (
        "Static3_Cottage Furniture $750 Offer_Statics_March2026",
        "Use client photos from Drive folder: drive.google.com/drive/folders/1zfqm-Rr5mYWZvi58Zh7sTHgpnodTwtxb\nStrongest image: farmhouse table or cottage living room with visible wood grain and warm light. Formats: 1:1, 4:5, 9:16.\nReference: tryatria.com/ad/m1417334119738520",
        "Create a home that feels relaxed, warm, and effortless.\n\nFor a limited time, get $750 OFF your first custom furniture order of $5,000 or more.\n\nEvery piece is built to your exact size and finish — solid Canadian hardwood, delivered across Ontario.\n\nBook a free consultation to get started.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "$750 OFF — Custom Cottage Furniture",
        "On orders of $5,000 or more. Limited time.",
        "Claim Offer",
        "https://go.parkroadfurniture.com/cottage-furniture",
        "Overlay: Headline — '$750 OFF When You Spend $5,000+' | Body — 'Create a home that feels relaxed, warm, and effortless' | Footer",
    ),
    (
        "Static4_Cottage Furniture $750 Offer_Statics_March2026",
        "Rich, warm-toned image. Dark walnut table or warm oak bedroom set. A candle, a book, a glass — something signals this space is lived in and loved. Moody warmth, not cold minimalism. Formats: 1:1, 4:5.\nReference: motionapp.com/…/inspo/ad/1351424240158359",
        "Natural materials. Thoughtful design. A space made for you.\n\nThe right furniture doesn't just fill a room. It changes how the room feels.\n\nPark Road builds custom pieces in solid Canadian hardwood — dining tables, sofas, bedroom sets — all made to order in London, Ontario.\n\n$750 OFF when you spend $5,000+.\n\n📍 761 Fanshawe Park Rd W, London   📞 (519) 660-0074",
        "Make Your Home Feel Warmer",
        "Solid Canadian hardwood. Built to order.",
        "Get Quote",
        "https://go.parkroadfurniture.com/cottage-furniture",
        "Overlay: Headline — 'Make Your Home Feel Warmer' | Body (stacked) — 'Natural materials / Thoughtful design / A space made for you' | '$750 OFF When You Spend $5,000+' | Footer",
    ),
]

# ── Audiences ────────────────────────────────────────────────────────────────
AUDIENCES = [
    (
        "Park Road — London Homeowners",
        "Custom Furniture (C1)",
        "London ON + 60 km",
        "30–65+",
        "Home furnishings, Interior design, Custom furniture, Home improvement, Bedroom furniture, Dining room",
        "Homeowners, Likely to move (next 3 months)",
        "Recent lead submitters (30d)",
        "Manual saved audience for Ad Set 2 of C1",
    ),
    (
        "Park Road — Ontario Cottage Buyers",
        "Cottage Furniture (C2)",
        "Ontario province-wide",
        "35–65+",
        "Cottage, Muskoka region, Vacation home, Farmhouse decor, Rustic home decor, Interior design",
        "Homeowners, Engaged shoppers",
        "Recent lead submitters (30d)",
        "Manual saved audience for Ad Set 2 of C2",
    ),
    (
        "Park Road — Website Retargeting 30d",
        "Both (separate retargeting ad set)",
        "Match campaign geo",
        "All ages",
        "N/A — Pixel-based (all website visitors)",
        "Anyone who visited go.parkroadfurniture.com in last 30 days",
        "Anyone who completed Lead event (already converted)",
        "Activate once website has 500+ monthly visitors. Budget: $10/day. Copy angle: 'Still thinking about it? $750 off expires soon.'",
    ),
    (
        "Park Road — CF Lookalike 1%",
        "Custom Furniture (C1)",
        "Canada",
        "All",
        "N/A — Lookalike based on CF - Lead (Meta) converters",
        "1% Lookalike of Lead event completers",
        "Existing converters",
        "Activate after 50+ CF - Lead (Meta) events. New ad set, compare CPL vs. interest targeting.",
    ),
    (
        "Park Road — Cottage Lookalike 1%",
        "Cottage Furniture (C2)",
        "Canada",
        "All",
        "N/A — Lookalike based on Cottage - Lead (Meta) converters",
        "1% Lookalike of Lead event completers",
        "Existing converters",
        "Activate after 50+ Cottage - Lead (Meta) events.",
    ),
]

# ── Pixel / Tracking ──────────────────────────────────────────────────────────
CONVERSIONS = [
    ("CF - Lead (Meta)",       "Lead",  "content_name = CF - Form Submission",       "Custom Furniture (C1)", "/custom-furniture/thank-you",       "7-day click, 1-day view"),
    ("Cottage - Lead (Meta)",  "Lead",  "content_name = Cottage - Form Submission",  "Cottage Furniture (C2)", "/cottage-furniture/thank-you",     "7-day click, 1-day view"),
]

PIXEL_PREREQS = [
    "Meta Pixel installed on go.parkroadfurniture.com (via GHL or direct code)",
    "Lead event firing on /custom-furniture/thank-you with content_name: 'CF - Form Submission'",
    "Lead event firing on /cottage-furniture/thank-you with content_name: 'Cottage - Form Submission'",
    "Verify in Events Manager — test each thank-you page and confirm Lead event appears",
    "Conversions API (CAPI) set up if possible — reduces iOS 14+ signal loss",
    "Custom conversion 'CF - Lead (Meta)' created in Events Manager",
    "Custom conversion 'Cottage - Lead (Meta)' created in Events Manager",
]

# ── Budget ────────────────────────────────────────────────────────────────────
BUDGET = [
    ("Custom Furniture",  "custom_furniture_advantageplus_london",  "$20/day", "~$600/month",  "Advantage+ audience, London + 60km"),
    ("Custom Furniture",  "custom_furniture_interest_london",       "$15/day", "~$450/month",  "Interest targeting, SW Ontario cities"),
    ("Cottage Furniture", "cottage_furniture_advantageplus_ontario","$25/day", "~$750/month",  "Advantage+ audience, Ontario province-wide"),
    ("Cottage Furniture", "cottage_furniture_interest_ontario",     "$20/day", "~$600/month",  "Interest targeting, Ontario province-wide"),
]

# ── Launch Checklist ─────────────────────────────────────────────────────────
LAUNCH = [
    ("Before Launch", "Meta Pixel installed on go.parkroadfurniture.com"),
    ("Before Launch", "Lead event verified on /custom-furniture/thank-you"),
    ("Before Launch", "Lead event verified on /cottage-furniture/thank-you"),
    ("Before Launch", "Custom conversions created: CF - Lead (Meta) and Cottage - Lead (Meta)"),
    ("Before Launch", "All 8 static images produced in 1:1, 4:5, and 9:16 formats"),
    ("Before Launch", "Client photos sourced from Drive or new shoot organized"),
    ("Before Launch", "Business Manager / Ad Account access confirmed for Hoski"),
    ("Before Launch", "Meta payment method confirmed active (was paused due to payment failure — confirm with Christopher)"),
    ("Launch Day",    "Create both campaigns in Meta Ads Manager"),
    ("Launch Day",    "Set up 2 ad sets per campaign with correct geo, audience, and budget"),
    ("Launch Day",    "Upload all 8 creatives with primary text, headlines, and CTAs"),
    ("Launch Day",    "Set conversion event: CF - Lead (Meta) for C1, Cottage - Lead (Meta) for C2"),
    ("Launch Day",    "Review targeting once more — then set both campaigns to Active"),
    ("Week 1–2",      "Check delivery — are campaigns exiting Learning Phase? (50+ optimization events needed)"),
    ("Week 1–2",      "Monitor CPL — if over $120, check creative quality, landing page, and audience overlap"),
    ("Week 1–2",      "Check ad frequency — if above 3.0 within 2 weeks, audience too small or budget too high"),
    ("Week 1–2",      "Pause weakest creative per ad set if clear winner emerges"),
    ("Day 30",        "Compare CPL: Meta vs. Google Ads (by campaign — CF vs. Cottage)"),
    ("Day 30",        "Identify winning ad set (Advantage+ vs. Interest) — consolidate budget to winner"),
    ("Day 30",        "If 50+ Lead events per campaign: build Lookalike audiences and launch new ad sets"),
    ("Day 30",        "If 500+ website visitors: activate retargeting ad set with urgency copy"),
    ("Day 30",        "Review winning creative angle — brief new creatives based on best performer"),
    ("Day 30",        "Increase budget 20% on any ad set with CPL under $60"),
]


# ════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
wb.remove(wb.active)


# ── Tab 1: Campaign Settings ─────────────────────────────────────────────────
ws = wb.create_sheet("Campaign Settings")
ws.sheet_view.showGridLines = False
set_widths(ws, [32, 50, 50])
freeze(ws, "A2")

header_row(ws, 1, ["Setting", "C1: Custom Furniture", "C2: Cottage Furniture"], C_DARK)
for i, (s, c1, c2) in enumerate(CAMP_SETTINGS, 2):
    bg = C_BLUE_LIGHT if i % 2 == 0 else C_GREEN_LIGHT
    wrap(ws, i, 1, s, fill=C_AMBER_LIGHT, bold=True)
    wrap(ws, i, 2, c1, fill=C_BLUE_LIGHT)
    wrap(ws, i, 3, c2, fill=C_GREEN_LIGHT)
    ws.row_dimensions[i].height = 30


# ── Tab 2: Ad Set Structure ──────────────────────────────────────────────────
ws = wb.create_sheet("Ad Set Structure")
ws.sheet_view.showGridLines = False
set_widths(ws, [22, 40, 22, 42, 10, 10, 55, 45])
freeze(ws, "A2")

cols = ["Campaign", "Ad Set Name", "Audience Type", "Geo", "Budget", "Age", "Targeting Details", "Placements"]
header_row(ws, 1, cols, C_DARK)

for i, row in enumerate(ADSETS, 2):
    camp = row[0]
    color = C_BLUE_LIGHT if camp == "Custom Furniture" else C_GREEN_LIGHT
    head  = C_BLUE_HEAD  if camp == "Custom Furniture" else C_GREEN_HEAD
    for c, val in enumerate(row, 1):
        wrap(ws, i, c, val, fill=color, bold=(c == 1), font_color="FFFFFF" if c == 1 else "000000")
        if c == 1:
            ws.cell(row=i, column=1).fill = hfill(head)
    ws.row_dimensions[i].height = 90


# ── Tab 3: CF Creatives ───────────────────────────────────────────────────────
ws = wb.create_sheet("C1 — CF Creatives")
ws.sheet_view.showGridLines = False
set_widths(ws, [45, 45, 55, 38, 35, 14, 48])
freeze(ws, "A2")

cols = ["File Name", "Visual Direction", "Primary Text (Body)", "Headline", "Description", "CTA", "Destination URL"]
header_row(ws, 1, cols, C_BLUE_HEAD)

for i, row in enumerate(CF_CREATIVES, 2):
    bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
    for c, val in enumerate(row, 1):
        wrap(ws, i, c, val, fill=bg)
    ws.row_dimensions[i].height = 130


# ── Tab 4: Cottage Creatives ──────────────────────────────────────────────────
ws = wb.create_sheet("C2 — Cottage Creatives")
ws.sheet_view.showGridLines = False
set_widths(ws, [45, 52, 55, 38, 35, 14, 48, 52])
freeze(ws, "A2")

cols = ["File Name", "Visual Direction", "Primary Text (Body)", "Headline", "Description", "CTA", "Destination URL", "Image Overlay Copy"]
header_row(ws, 1, cols, C_GREEN_HEAD)

for i, row in enumerate(C2_CREATIVES, 2):
    bg = C_GREEN_LIGHT if i % 2 == 0 else C_WHITE
    for c, val in enumerate(row, 1):
        wrap(ws, i, c, val, fill=bg)
    ws.row_dimensions[i].height = 150


# ── Tab 5: Pixel + Tracking ───────────────────────────────────────────────────
ws = wb.create_sheet("Pixel + Tracking")
ws.sheet_view.showGridLines = False
set_widths(ws, [28, 16, 42, 24, 42, 28])
freeze(ws, "A2")

# Prerequisites
section_title(ws, 1, "Prerequisites — Before Launch", 3, C_AMBER_HEAD)
header_row(ws, 2, ["#", "Status", "Task"], C_AMBER_HEAD)
for i, task in enumerate(PIXEL_PREREQS, 3):
    wrap(ws, i, 1, str(i - 2), fill=C_AMBER_LIGHT, bold=True, h_align="center")
    wrap(ws, i, 2, "[ ]", fill=C_AMBER_LIGHT, h_align="center")
    wrap(ws, i, 3, task, fill=C_AMBER_LIGHT)
    ws.row_dimensions[i].height = 25

# Conversions table
r = len(PIXEL_PREREQS) + 5
section_title(ws, r, "Custom Conversions to Create in Events Manager", 6, C_AMBER_HEAD)
r += 1
header_row(ws, r, ["Conversion Name", "Event", "Filter Parameter", "Campaign", "Thank-You URL", "Attribution Window"], C_AMBER_HEAD)
r += 1
for i, row in enumerate(CONVERSIONS):
    bg = C_AMBER_LIGHT if i % 2 == 0 else C_WHITE
    for c, val in enumerate(row, 1):
        wrap(ws, r, c, val, fill=bg)
    ws.row_dimensions[r].height = 25
    r += 1


# ── Tab 6: Audience Library ───────────────────────────────────────────────────
ws = wb.create_sheet("Audience Library")
ws.sheet_view.showGridLines = False
set_widths(ws, [35, 22, 22, 12, 52, 45, 35, 55])
freeze(ws, "A2")

cols = ["Audience Name", "Campaign", "Geo", "Age", "Interests", "Behaviors / Source", "Exclusions", "Notes"]
header_row(ws, 1, cols, C_PURPLE_HEAD)

for i, row in enumerate(AUDIENCES, 2):
    bg = C_PURPLE_LIGHT if i % 2 == 0 else C_WHITE
    for c, val in enumerate(row, 1):
        wrap(ws, i, c, val, fill=bg)
    ws.row_dimensions[i].height = 80


# ── Tab 7: Budget + Launch ────────────────────────────────────────────────────
ws = wb.create_sheet("Budget + Launch")
ws.sheet_view.showGridLines = False
set_widths(ws, [22, 44, 12, 16, 48])
freeze(ws, "A2")

# Budget
section_title(ws, 1, "Budget Breakdown — Launch Phase", 5, C_TEAL_HEAD)
header_row(ws, 2, ["Campaign", "Ad Set Name", "Daily Budget", "Monthly Estimate", "Notes"], C_TEAL_HEAD)
for i, row in enumerate(BUDGET, 3):
    camp = row[0]
    bg = C_BLUE_LIGHT if camp == "Custom Furniture" else C_GREEN_LIGHT
    for c, val in enumerate(row, 1):
        wrap(ws, i, c, val, fill=bg)
    ws.row_dimensions[i].height = 22

# Totals row
r = len(BUDGET) + 3
wrap(ws, r, 1, "TOTAL — Launch Phase", fill=C_TEAL_HEAD, bold=True, font_color="FFFFFF")
wrap(ws, r, 2, "", fill=C_TEAL_HEAD)
wrap(ws, r, 3, "$80/day", fill=C_TEAL_HEAD, bold=True, font_color="FFFFFF")
wrap(ws, r, 4, "~$2,400/month", fill=C_TEAL_HEAD, bold=True, font_color="FFFFFF")
wrap(ws, r, 5, "CPL target: $50–$80/lead. 1 closed sale/month covers full Meta spend at $5,000+ AOV.", fill=C_TEAL_LIGHT)
ws.row_dimensions[r].height = 30

# Scaling rule
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
wrap(ws, r, 1,
     "Scaling rule: After 14 days, pause the weaker ad set per campaign (higher CPL). Do not increase any ad set budget by more than 20% in a 7-day window — larger increases reset the learning phase.",
     fill=C_TEAL_LIGHT)
ws.row_dimensions[r].height = 35

# Launch Checklist
r += 3
section_title(ws, r, "Launch Checklist", 5, C_TEAL_HEAD)
r += 1
header_row(ws, r, ["Phase", "Status", "Task", "", ""], C_TEAL_HEAD)
r += 1
current_phase = ""
for phase, task in LAUNCH:
    if phase != current_phase:
        current_phase = phase
    bg = C_TEAL_LIGHT if current_phase in ("Before Launch", "Day 30") else C_WHITE
    wrap(ws, r, 1, phase, fill=bg, bold=True)
    wrap(ws, r, 2, "[ ]", fill=bg, h_align="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    wrap(ws, r, 3, task, fill=bg)
    ws.row_dimensions[r].height = 22
    r += 1


# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
